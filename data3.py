#!/usr/bin/python3
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def read_data(input_file):
    data = []
    with open(input_file, "r") as file:
        for line in file:
            name, surname, age, profession = line.strip().split(",")
            data.append((name, surname, int(age), profession))
    return data

def sort_data(data, sort_option):
    if sort_option == "n":
        return sorted(data, key=lambda x: x[0])
    elif sort_option == "s":
        return sorted(data, key=lambda x: x[1])
    elif sort_option == "a":
        return sorted(data, key=lambda x: x[2])
    elif sort_option == "p":
        return sorted(data, key=lambda x: x[3])
    else:
        return data

def create_excel(data, output_file):
    df = pd.DataFrame(data, columns=["Name", "Surname", "Age", "Profession"])

    wb = Workbook()
    ws = wb.active
    for index, row in df.iterrows():
        ws.append(row.tolist())

    header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = yellow_fill

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[2].value > 20:
            for cell in row:
                cell.fill = green_fill

    wb.save(output_file)
    print(f"Excel file '{output_file}' created successfully.")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--input-file", required=True, help="Input data file name")
    parser.add_argument("-o", "--output-file", required=True, help="Output Excel file name")
    parser.add_argument("-s", "--sort-option", choices=["n", "s", "a", "p"], help="Sorting option")
    args = parser.parse_args()

    data = read_data(args.input_file)
    sorted_data = sort_data(data, args.sort_option)
    create_excel(sorted_data, args.output_file)

if __name__ == "__main__":
    main()
