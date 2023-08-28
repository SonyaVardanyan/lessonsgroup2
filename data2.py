#!/usr/bin/python3
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument("-f", "--input-file", required=True, help="Input data file name")
parser.add_argument("-o", "--output-file", required=True, help="Output Excel file name")
parser.add_argument("-s", "--sort-option", choices=["n", "s", "a", "p"], help="Sorting option")
args = parser.parse_args()

# Read data from input file
data = []
with open(args.input_file, "r") as file:
    for line in file:
        name, surname, age, profession = line.strip().split(",")
        data.append((name, surname, int(age), profession))

# Apply sorting if specified
if args.sort_option == "n":
    data.sort(key=lambda x: x[0])
elif args.sort_option == "s":
    data.sort(key=lambda x: x[1])
elif args.sort_option == "a":
    data.sort(key=lambda x: x[2])
elif args.sort_option == "p":
    data.sort(key=lambda x: x[3])

# Create a Pandas DataFrame
df = pd.DataFrame(data, columns=["Name", "Surname", "Age", "Profession"])

# Create an Excel writer and add the DataFrame
wb = Workbook()
ws = wb.active
for index, row in df.iterrows():
    ws.append(row.tolist())


# Apply formatting
header_font = Font(bold=True)
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Apply formatting to header
for cell in ws[1]:
    cell.font = header_font
    cell.fill = yellow_fill

# Apply green color fill to rows with age > 20
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    if row[2].value > 20:
        for cell in row:
            cell.fill = green_fill

# Save the Excel file
wb.save(args.output_file)
print(f"Excel file '{args.output_file}' created successfully.")
